import requests  # Solicitações HTTP
from bs4 import BeautifulSoup  # Analisar e extrair dados de HTML
from selenium import webdriver  # Automação de navegador
import openpyxl  # Manipulação de arquivos Excel
import os.path  # Maipulação de arquivos e diretorios
from datetime import datetime  # Manipulação de data e hora

# Definindo constantes
GOOGLE_URL = "https://www.google.com/search?q=dolar"
DOLAR_CLASS = "DFlfde"
PRECISION_ATTR = {"data-precision": "2"}
WORKSHEET_NAME = "cotacao_dolar"

# Verifica se o arquivo já existe
if not os.path.isfile("cotacao_dolar.xlsx"):
    # Cria uma nova planilha vazia
    workbook = openpyxl.Workbook()
    workbook.save("cotacao_dolar.xlsx")

# Iniciando o driver do navegador
driver = webdriver.Chrome()

try:
    # Acessando a URL do Google
    driver.get(GOOGLE_URL)

    # Obtendo o HTML da página do Google
    html = driver.page_source

finally:
    # Fechando o navegador, mesmo em caso de erro
    driver.quit()

# Utilizando o BeautifulSoup para fazer o web scraping da página do Google
soup = BeautifulSoup(html, "html.parser")

# Encontrando a tag que contém a cotação do dólar
dolar = soup.find("span", {"class": DOLAR_CLASS, **PRECISION_ATTR})

if dolar is not None:
    # Obtendo o valor da cotação do dólar
    valor_dolar = dolar.text

    # Abrindo a planilha do Excel
    workbook = openpyxl.load_workbook("cotacao_dolar.xlsx")

    # Selecionando a aba de destino
    worksheet = workbook["Sheet"]

    # Encontrando a próxima linha vazia na planilha
    next_row = worksheet.max_row + 1

    # Obtendo a data e hora atual
    agora = datetime.now().strftime("%d/%m/%Y às %H:%M")

    # Inserindo a data e hora da consulta na planilha
    worksheet.cell(row=next_row, column=1).value = f"Cotação do dólar em {agora}"

    # Inserindo o valor da cotação do dólar na planilha
    worksheet.cell(row=next_row, column=2).value = valor_dolar

    # Salvando a planilha
    workbook.save("cotacao_dolar.xlsx")
